from datetime import date
import pandas as pd
from models import engine
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

'''
Calculate the monthly AUM in an excel file, from the fund money or the leveraged AUM
'''

def get_monthly_aum(start_date, end_date, is_leveraged):
    if is_leveraged:
        my_type = 'leveraged'
        multiple = 1000000
        title = 'Leveraged'
    else:
        my_type = 'Fund'
        multiple = 1
        title = 'Fund'

    my_sql = f"""SELECT entry_date,amount*{multiple} as amount FROM aum WHERE type='{my_type}' AND fund_id=4
    AND entry_date>='{start_date}' and entry_date<='{end_date}' order by entry_date;"""
    df_aum = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'], index_col='entry_date')

    file_name = f'Excel/{title} Monthly AUM.xlsx'

    excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')
    df_aum.to_excel(excel_writer, sheet_name='AUM', startrow=0, index=True, header=True)
    excel_writer._save()

    workbook = load_workbook(file_name)
    sheet = workbook['AUM']
    sheet.freeze_panes = sheet.cell(row=2, column=1)

    for cell in sheet['A']:
        cell.number_format = 'DD/MM/YYYY'
    for cell in sheet['B']:
        cell.number_format = "#,##0"

    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.4  # Adding some buffer and a conversion factor
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    fill = PatternFill(start_color='66b8cc', end_color='66b8cc', fill_type='solid')

    for col in range(1, 3):
        cell = sheet.cell(row=1, column=col)
        cell.fill = fill

    workbook.save(file_name)


if __name__ == '__main__':
    start_date = date(2019, 3, 29)
    end_date = date(2023, 12, 31)

    is_leveraged = False

    get_monthly_aum(start_date, end_date, is_leveraged)