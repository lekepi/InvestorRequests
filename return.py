import pandas as pd
from models import engine
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers


def get_return_data(calcul_type, classification_list, numerator_list, denominator, start_date, end_date,
                    is_leveraged, is_close_pnl, period):

    my_sql = f"""SELECT T1.entry_date,T2.ticker,T2.prod_type,T4.name as Country,T4.continent as Continent,
            T5.name as Sector,T1.pnl_usd,T1.alpha_usd,mkt_value_usd as notional_usd
            FROM position T1 JOIN product T2 on T1.product_id=T2.id
            LEFT JOIN exchange T3 on T2.exchange_id=T3.id LEFT JOIN country T4 on T3.country_id=T4.id
            LEFT JOIN industry_group_gics T5 on T2.industry_group_gics_id=T5.id WHERE parent_fund_id=1
            and entry_date>='{start_date}' and entry_date<='{end_date}' and (T2.prod_type='Cash' or 
            T2.ticker in ('ES1 CME', 'SXO1 EUX')) order by entry_date,T2.ticker"""

    df = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
    df['pnl_usd'] = df['pnl_usd'].fillna(0)
    df['alpha_usd'] = df['alpha_usd'].fillna(0)
    df['month_year'] = df['entry_date'].dt.strftime('%Y-%m')

    if is_close_pnl:
        my_sql = f"""SELECT T1.trade_date as entry_date,T2.prod_type,T4.name as Country,T4.continent as Continent,
            T5.name as Sector,CASE WHEN T6.generic_future IS NOT NULL THEN T6.generic_future ELSE T2.ticker END as tickerx,sum(T1.pnl_close) as pnl_close FROM trade T1 
            JOIN product T2 on T1.product_id=T2.id LEFT JOIN exchange T3 on T2.exchange_id=T3.id LEFT JOIN country T4 on T3.country_id=T4.id
            LEFT JOIN industry_group_gics T5 on T2.industry_group_gics_id=T5.id
            LEFT JOIN security T6 on T2.security_id=T6.id WHERE T1.parent_fund_id=1 and trade_date>='2019-04-01'
            and trade_date>='{start_date}' and trade_date<='{end_date}' and (T2.prod_type in ('Cash', 'Future') or T6.generic_future in 
            ('ES1 CME', 'SXO1 EUX')) GROUP BY trade_date,tickerx,T2.prod_type,Country,continent,sector order by T1.trade_date,tickerx"""
        df_close = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        df_close = df_close.rename(columns={'tickerx': 'ticker'})
        df = pd.merge(df, df_close, on=['entry_date', 'ticker', 'prod_type', 'Country', 'Continent', 'Sector'], how='outer')

    df['pnl_close'] = df['pnl_close'].fillna(0)
    df['pnl_usd'] = df['pnl_usd'].fillna(0)
    df['alpha_usd'] = df['alpha_usd'].fillna(0)
    df['notional_usd'] = df['notional_usd'].fillna(0)

    df['month_year'] = df['entry_date'].dt.strftime('%Y-%m')
    df['week'] = df['entry_date'].dt.to_period('W').apply(lambda r: r.start_time)
    df['year'] = df['entry_date'].dt.year
    df['quarter'] = df['entry_date'].dt.to_period('Q')

    if is_close_pnl:
        df['pnl_usd'] = df['pnl_usd'] + df['pnl_close']
        df['alpha_usd'] = df['alpha_usd'] + df['pnl_close']

     # sort by ticker, entry_date
    df = df.sort_values(by=['ticker', 'entry_date'])

    if 'MarketCap' in classification_list:
        # get list of start date of the month
        my_sql = f"""SELECT MIN(entry_date) AS first_date FROM position WHERE entry_date>='{start_date}' 
            and entry_date<='{end_date}' GROUP BY YEAR(entry_date), MONTH(entry_date) order by entry_date"""
        df_date = pd.read_sql(my_sql, con=engine, parse_dates=['first_date'])
        first_date_list = df_date['first_date'].tolist()
        first_date_str = ','.join([f"'{first_date}'" for first_date in first_date_list])

        my_sql = f"""SELECT T1.entry_date,T2.ticker,CASE WHEN T3.market_cap<3000 then '0-3Bn' WHEN T3.market_cap<10000 THEN '3-10Bn' else '>10Bn' END as marketCap
        FROM position T1 JOIN product T2 on T1.product_id=T2.id JOIN product_market_cap T3 on T1.product_id=T3.product_id and T1.entry_date=T3.entry_date and T3.type='Monthly'
        WHERE T1.entry_date in ({first_date_str})  and T2.prod_type='Cash' and T1.parent_fund_id=1
        GROUP by T1.entry_date,T2.ticker;"""
        df_market_cap = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        df = df.merge(df_market_cap, on=['entry_date', 'ticker'], how='left')
        df['MarketCap'] = df.groupby('ticker')['marketCap'].fillna(method='ffill')
        df['MarketCap'] = df.apply(lambda x: 'Index' if x['ticker'] in ('ES1 CME', 'SXO1 EUX') else x['MarketCap'], axis=1)

    # get the denominator: long, short, net, gross, nav
    if denominator == 'Long':
        df_denominator = df[df['notional_usd'] >= 0]
        # group by entry_date not month_year
        df_denominator = df_denominator.groupby('entry_date').agg({'notional_usd': 'sum'}).reset_index()
        df = df.merge(df_denominator, on='entry_date', how='left')

    elif denominator == 'Nav':
        my_sql = f"""Select entry_date,amount*1000000 as denominator from aum where entry_date>='2019-04-01' 
                and entry_date>='{start_date}' and entry_date<='{end_date}' and type='leveraged' and fund_id=4 order by entry_date"""
        df_denominator = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        df_denominator['month_year'] = df_denominator['entry_date'].dt.strftime('%Y-%m')
        # remove entry_date
        df_denominator = df_denominator.drop(columns=['entry_date'])
        df = df.merge(df_denominator, on='month_year', how='left')

    df['PnL %'] = df['pnl_usd'] / df['denominator']
    df['Alpha %'] = df['alpha_usd'] / df['denominator']

    if is_leveraged and denominator == 'Nav':
        df['PnL %'] = df['PnL %'] * 2
        df['Alpha %'] = df['Alpha %'] * 2

    df['Sector'] = df.apply(lambda x: 'Index' if x['ticker'] in ('ES1 CME', 'SXO1 EUX') else x['Sector'], axis=1)
    df['Country'] = df.apply(lambda x: 'S&P500' if x['ticker'] in ('ES1 CME') else x['Country'], axis=1)
    df['Country'] = df.apply(lambda x: 'Stoxx 600' if x['ticker'] in ('SXO1 EUX') else x['Country'], axis=1)

    df.sort_values(by='entry_date', ascending=False, inplace=True)

    if period == 'Daily':
        group_period = 'entry_date'
    elif period == 'Weekly':
        group_period = 'week'
    elif period == 'Monthly':
        group_period = 'month_year'
    elif period == 'Quarterly':
        group_period = 'quarter'
    elif period == 'Yearly':
        group_period = 'year'

    df = df.groupby(['ticker', 'prod_type', 'Country', 'Continent', 'Sector', 'MarketCap', group_period]).agg({
        'pnl_usd': 'sum',
        'alpha_usd': 'sum',
        'notional_usd': 'sum',
        'pnl_close': 'sum',
        'PnL %': 'sum',
        'Alpha %': 'sum'
        }).reset_index()

    result_list = []
    for classification in classification_list:
        df_result = pd.DataFrame()
        numerator_len_list = []
        for numerator in numerator_list:
            df_temp = df.copy()
            if numerator == 'Long':
                df_temp = df_temp[df_temp['notional_usd'] >= 0]
            elif numerator == 'Short':
                df_temp = df_temp[df_temp['notional_usd'] < 0]

            if (numerator == 'Long' and denominator == 'Short') or (numerator == 'Short' and denominator == 'Long'):
                df_temp['Exposure'] = df_temp['Exposure'] * -1

            df_temp = df_temp.groupby([group_period, classification]).agg({f'{calcul_type}': 'sum'}).reset_index()
            df_pivot = df_temp.pivot(index=group_period, columns=classification, values=calcul_type).fillna(0)
            if 'Index' in df_pivot.columns:
                df_pivot = df_pivot[[col for col in df_pivot.columns if col != 'Index'] + ['Index']]
            if 'S&P500' in df_pivot.columns:
                df_pivot = df_pivot[[col for col in df_pivot.columns if col != 'S&P500'] + ['S&P500']]
            if 'Stoxx 600' in df_pivot.columns:
                df_pivot = df_pivot[[col for col in df_pivot.columns if col != 'Stoxx 600'] + ['Stoxx 600']]
            if len(numerator_list) > 1:
                # add numerator to the column name
                df_pivot.columns = [f'{col} {numerator}' for col in df_pivot.columns]
            numerator_len_list.append(len(df_pivot.columns))
            if df_result.empty:
                df_result = df_pivot
            else:
                df_result = df_result.merge(df_pivot, on=group_period, how='outer')
        df_result = df_result.sort_values(by=group_period, ascending=False)
        result_list.append([classification, df_result, numerator_len_list])

    if len(numerator_list) == 1:
        numerator_str = numerator_list[0] + ' '
    else:
        numerator_str = ' '

    file_name = f'Excel/{period} {numerator_str}{calcul_type} Over {denominator}.xlsx'
    with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
        for result in result_list:
            result[1].to_excel(writer, sheet_name=result[0], header=True, index=True)

    # complete formatting

    fill1 = PatternFill(start_color='66b8cc', end_color='66b8cc', fill_type='solid')
    fill2 = PatternFill(start_color='f08c30', end_color='f08c30', fill_type='solid')

    workbook = load_workbook(file_name)
    for result in result_list:
        worksheet = workbook[result[0]]
        numerator_len_list = result[2]

        df_temp = result[1]  # get the df
        num_cols = len(df_temp.columns)
        # header color

        # for col_num in range(1, num_cols + 2):
        #    worksheet.cell(row=1, column=col_num).fill = fill1
        current_fill = fill1
        col_index = 1
        for num_cols_same_color in numerator_len_list:
            for _ in range(num_cols_same_color):
                worksheet.cell(row=1, column=col_index+1).fill = current_fill
                col_index += 1
            # Switch fill
            current_fill = fill2 if current_fill == fill1 else fill1

            # Format columns 2 to the last as percentages
        # % format
        for col_num in range(2, num_cols + 2):
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            for cell in worksheet[column_letter]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

        for i, col in enumerate(worksheet.columns):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = ((max_length+2) * 1.2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Add freeze pane to the second row
        worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    workbook.save(file_name)


if __name__ == '__main__':

    # calcul_type = 'Exposure', 'PnL', 'Alpha'
    # classification_list = ['Continent', 'Country', 'Sector', 'MarketCap']
    # numerator_list = ['Long'], ['Short'], ['Net'], ['Gross']
    # denominator = 'Long', 'Nav', 'Gross', 'Net'
    # period = 'Daily', 'Weekly', 'Monthly', 'Quarterly', 'Yearly'

    start_date = date(2019, 4, 1)
    end_date = date(2024, 2, 29)

    # pnl vs Nav
    get_return_data(calcul_type='PnL %',  # 'Alpha %'
                    classification_list=['Continent', 'Sector', 'MarketCap'],
                    numerator_list=['Net', 'Long', 'Short'],
                    denominator='Nav',
                    start_date=start_date,
                    end_date=end_date,
                    is_leveraged=True,
                    is_close_pnl=True,
                    period='Monthly')

