import pandas as pd
from models import engine
from datetime import date, timedelta
from utils import simple_email
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font



"""
Get the turnover by capital: for trade, we take all the net daily trade for all asset classes
and divide it by the gross exposure (sum of all the absolute value Long+Short of the market value of the positions)
per year

Get the turnover by name 
"""


def last_weekday_of_year(year):
    last_day = date(year, 12, 31)  # Set initial date to the last day of the given year
    while last_day.weekday() >= 5:  # 5 and 6 represent Saturday and Sunday
        last_day -= timedelta(days=1)
    return last_day


def turnover_capital(my_type='Gross Exposure', is_addition_redemption=True):

    my_sql = """SELECT entry_date, 1000000*amount as aum FROM aum WHERE type = 'leveraged' and entry_date>='2019-04-01';"""
    df_aum = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    my_sql = """SELECT trade_date,T2.ticker,ABS(SUM(T1.notional_usd)) AS trade_usd FROM trade T1 
    JOIN product T2 ON T1.product_id = T2.id WHERE T2.still_active=1 and trade_date>='2019-04-01'
    and T1.parent_fund_id=1 GROUP BY trade_date,ticker ORDER BY trade_date,ticker;"""

    df_trade = pd.read_sql(my_sql, con=engine, parse_dates=['trade_date'])
    # group by trade_date
    df_trade = df_trade.groupby('trade_date')['trade_usd'].sum().reset_index()
    # rename trade_date to entry_date
    df_trade = df_trade.rename(columns={'trade_date': 'entry_date'})

    if my_type == 'Gross Exposure':
        my_sql = """SELECT entry_date, sum(abs(mkt_value_usd)) as gross_usd FROM position T1 GROUP BY entry_date;"""
        df_gross = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        # merge df_trade and df_gross
        df = pd.merge(df_trade, df_gross, on='entry_date', how='left')
        df['turnover'] = df['trade_usd'] / df['gross_usd']

    else:  # Levereged AUM
        # merge df_trade and df_aum
        df = pd.merge(df_trade, df_aum, on='entry_date', how='left')
        # sort by entry_date
        df = df.sort_values(by=['entry_date'])
        # fill na with previous value
        df['aum'] = df['aum'].fillna(method='ffill')
        df['turnover'] = df['trade_usd'] / df['aum']

    # group by year
    df_year = df.groupby(df['entry_date'].dt.year)['turnover'].sum().reset_index()
    df_year = df_year.rename(columns={'entry_date': 'year'})

    if is_addition_redemption:
        # adjustment for additions/redemptions
        df_aum['aum_change'] = df_aum['aum'].diff().shift(-1).abs()
        df_aum['turnover_adj'] = -df_aum['aum_change'] / df_aum['aum']
        # group by year
        df_year_aum = df_aum.groupby(df_aum['entry_date'].dt.year)['turnover_adj'].sum().reset_index()
        df_year_aum = df_year_aum.rename(columns={'entry_date': 'year'})
        df_year = pd.merge(df_year, df_year_aum, on='year', how='left')
        df_year['turnover'] = df_year['turnover'] + df_year['turnover_adj']
        df_year = df_year.drop(columns=['turnover_adj'])
    # remove current year
    current_year = date.today().year
    df_year = df_year[df_year['year'] != current_year]

    # round to 2 decimals
    df_year['turnover'] = df_year['turnover'].apply(lambda x: round(x, 2))
    # sort by year desc
    df_year = df_year.sort_values(by=['year'], ascending=False)

    # send email
    # html = df_year.to_html(index=False)
    # subject = f'Turnover by capital - {my_type} Add/Redemp={is_addition_redemption}'
    # simple_email(subject, '', 'olivier@ananda-am.com', html=html)

    return df_year


def turnover_name(name_number):

    first_date = date(2019, 4, 1)
    current_year = date.today().year

    date_list = [first_date]

    for year in range(2019, current_year):
        last_day = last_weekday_of_year(year)
        date_list.append(last_day)

    df_result = pd.DataFrame(columns=['year', 'turnover'])

    df_old = pd.DataFrame()
    for index, my_date in enumerate(date_list):

        my_sql = f"""SELECT T2.ticker,mkt_value_usd FROM position T1 JOIN product T2 on T1.product_id=T2.id 
        WHERE parent_fund_id=1 and T2.prod_type='Cash' and entry_date='{my_date}' and mkt_value_usd>0 
        order by mkt_value_usd desc;"""
        df_new = pd.read_sql(my_sql, con=engine)

        if not df_old.empty:
            df_merge = pd.merge(df_old, df_new, on='ticker', how='left')
            # keep name_number first rows
            df_merge = df_merge.head(name_number)
            name_number_temp = df_merge['ticker'].count()
            # count the number of row with mkt_value_usd_y not null
            count = df_merge['mkt_value_usd_y'].count()
            turnover = 1 - count / name_number_temp
            df_result = df_result._append({'year': my_date.year, 'turnover': turnover}, ignore_index=True)
            # change turnover to string pct
            # reformat
        df_old = df_new

    df_result['turnover'] = df_result['turnover'].apply(lambda x: "{:.2%}".format(x))
    # year into int
    df_result['year'] = df_result['year'].astype(int).astype(str)
    # sort by year desc
    df_result = df_result.sort_values(by=['year'], ascending=False)

    # send email
    # html = df_result.to_html(index=False)
    # subject = f'Turnover by name - TOP {name_number} names'
    # simple_email(subject, '', 'olivier@ananda-am.com', html=html)

    return df_result


def get_turnover():
    name_number = 30
    my_type = "Gross Exposure"  # "AUM Leveraged"
    is_addition_redemption = True

    df_capital = turnover_capital(my_type, is_addition_redemption)
    df_name = turnover_name(name_number)

    filename = 'Excel/Turnover.xlsx'
    excel_writer = pd.ExcelWriter(filename, engine='openpyxl')
    df_capital.to_excel(excel_writer, sheet_name='Turnover by Capital', startrow=0, index=False, header=True)
    df_name.to_excel(excel_writer, sheet_name='Turnover by Name', startrow=0, index=False, header=True)
    excel_writer._save()

    workbook = load_workbook(filename)

    fill = PatternFill(start_color='66b8cc', end_color='66b8cc', fill_type='solid')

    note_list = ["* Turnover by capital is calculated as the net daily volume traded excluding the trades linked to investors redemptions and subscriptions over the gross portfolio exposure.",
                 f"* Turnover by name represents the % of top 30 long core names that are absent from the Long Portfolio after one year."]
    size_list = [len(df_capital), len(df_name)]

    sheet_list = workbook.sheetnames
    for index, sheet_name in enumerate(sheet_list):
        sheet = workbook[sheet_name]
        note = note_list[index]
        size = size_list[index]
        for col in range(1, 3):
            cell = sheet.cell(row=1, column=col)
            cell.fill = fill
        cell = sheet.cell(row=size+3, column=1)
        cell.value = note
        cell.font = Font(size=8)
    workbook.save(filename)


if __name__ == '__main__':

    get_turnover()




