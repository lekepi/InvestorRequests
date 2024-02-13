import pandas as pd
from models import engine, config_class
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

"""
Get the top 5 contributors/detractors in cash equity in absolute pnl (% of AUM) per month and per year
"""


def get_top_contributors():
    my_sql = """SELECT T2.ticker,T1.entry_date,T1.pnl_usd FROM position T1
JOIN Product T2 ON T1.product_id = T2.id WHERE prod_type = 'Cash' AND entry_date>='2019-04-01' AND parent_fund_id=1
order by entry_date"""

    df_position = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    my_sql = """SELECT entry_date,amount*1000000 as amount FROM aum WHERE type='leveraged';"""
    df_aum = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    # merge
    df_temp = pd.merge(df_position, df_aum, on='entry_date', how='left')
    # fill with previous value
    df_temp['amount'] = df_temp['amount'].fillna(method='ffill')
    df_temp['perf'] = df_temp['pnl_usd'] / df_temp['amount']

    df_temp['month_year'] = df_temp['entry_date'].dt.strftime('%Y-%m')
    df = df_temp.groupby(['month_year', 'ticker'])['perf'].sum().reset_index()

    # sort by month_year and perf
    df = df.sort_values(by=['month_year', 'perf'], ascending=[True, False])

    # get all unique month_year
    month_year = df['month_year'].unique()
    # get top 5 contributors
    df_top_month = pd.DataFrame()
    df_bottom_month = pd.DataFrame()
    for my_month_year in month_year:
        df_current_month = df[df['month_year'] == my_month_year]

        df_top_5 = df_current_month.head(5)
        df_top_month[my_month_year] = df_top_5['ticker'].values

        df_bottom_5 = df_current_month.tail(5)
        df_bottom_month[my_month_year] = df_bottom_5['ticker'].values

    # df_top_month.to_excel('Excel/Top 5 contributors by month.xlsx', index=False)
    # df_bottom_month.to_excel('Excel/Top 5 detractors by month.xlsx', index=False)

    df_temp['year'] = df_temp['entry_date'].dt.strftime('%Y')
    df = df_temp.groupby(['year', 'ticker'])['perf'].sum().reset_index()
    df = df.sort_values(by=['year', 'perf'], ascending=[True, False])
    df_top_year = pd.DataFrame()
    df_bottom_year = pd.DataFrame()

    year = df['year'].unique()
    for my_year in year:
        df_top_5 = df[df['year'] == my_year].head(5)
        df_top_year[my_year] = df_top_5['ticker'].values

        df_bottom_5 = df[df['year'] == my_year].tail(5)
        df_bottom_year[my_year] = df_bottom_5['ticker'].values

    # df_top_year.to_excel('Excel/Top 5 contributors by year.xlsx', index=False)
    # df_bottom_year.to_excel('Excel/Top 5 detractors by year.xlsx', index=False)

    excel_writer = pd.ExcelWriter('Excel/Contributors-Detractors.xlsx', engine='openpyxl')
    df_top_month.to_excel(excel_writer, sheet_name='Sheet1', startrow=1, index=False, header=True)
    df_bottom_month.to_excel(excel_writer, sheet_name='Sheet1', startrow=10, index=False, header=True)
    df_top_year.to_excel(excel_writer, sheet_name='Sheet1', startrow=19, index=False, header=True)
    df_bottom_year.to_excel(excel_writer, sheet_name='Sheet1', startrow=28, index=False, header=True)

    excel_writer._save()

    workbook = load_workbook('Excel/Contributors-Detractors.xlsx')
    sheet = workbook.active

    # Merge cells and apply styles

    format_table = [
        ['A1:C1', 'A1', 'Top 5 Contributors by month', "66b8cc"],
        ['A10:C10', 'A10', 'Top 5 Detractors by month', "eb7373"],
        ['A19:C19', 'A19', 'Top 5 Contributors by year', "66b8cc"],
        ['A28:C28', 'A28', 'Top 5 Detractors by year', "eb7373"]
        ]

    for format_list in format_table:
        sheet.merge_cells(format_list[0])
        merged_cell = sheet[format_list[1]]
        merged_cell.value = format_list[2]
        merged_cell.font = Font(bold=True)
        merged_cell.fill = PatternFill(start_color=format_list[3], end_color=format_list[3], fill_type="solid")

    workbook.save('Excel/Contributors-Detractors.xlsx')


if __name__ == '__main__':
    get_top_contributors()
    pass

