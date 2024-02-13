import pandas as pd
from models import engine, session, Product, ExecFee, TaskChecker
from openpyxl import load_workbook
from openpyxl.styles import numbers, PatternFill
from openpyxl.utils import get_column_letter

"""
Get the biggest position on single stock (long and Short) in % of the AUM(fund) for each day
13% since inception for long, 8% since 2022...

Then get the average long and short position % of AUM_fund for the core position (top30) since inception
"""

def get_position_size():

    my_sql = "SELECT entry_date,long_usd as long_usd FROM alpha_summary WHERE parent_fund_id=1 order by entry_date;"
    df_long = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    my_sql = "SELECT entry_date,amount*1000000 as nav_usd FROM aum WHERE type='leveraged' and entry_date>='2019-04-01';"
    df_nav = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    my_sql = """SELECT entry_date,T2.ticker,mkt_value_usd FROM position T1 JOIN product T2 on T1.product_id=T2.id WHERE prod_type='Cash' and parent_fund_id=1
    and entry_date>='2019-04-01' and mkt_value_usd is Not NULL order by entry_date;"""
    df_position = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
    df_position = pd.merge(df_position, df_long, on='entry_date', how='left')
    df_position = pd.merge(df_position, df_nav, on='entry_date', how='left')
    # fill the missing value with the previous value
    df_position['nav_usd'] = df_position['nav_usd'].fillna(method='ffill')

    df_position['pos % of Long'] = df_position['mkt_value_usd'] / df_position['long_usd']
    df_position['pos % of NAV'] = df_position['mkt_value_usd'] / df_position['nav_usd']

    df_position_long = df_position[df_position['mkt_value_usd'] > 0]
    df_position_short = df_position[df_position['mkt_value_usd'] < 0]
    df_position_long = df_position_long.sort_values(by='pos % of NAV', ascending=False)
    df_position_short = df_position_short.sort_values(by='pos % of NAV', ascending=True)

    count_list = [20, 30, 40]
    df_avg_size = pd.DataFrame(columns=['count', 'Avg Long Pos vs Long', 'Avg Short Pos vs Long',
                                        'Avg Long Pos vs NAV - Low Vol', 'Avg Short Pos vs NAV - Low Vol',
                                        'Avg Long Pos vs NAV - Alto', 'Avg Short Pos vs NAV - Alto'])

    for count in count_list:
        df_long_avg = df_position_long.sort_values(by=['entry_date', 'pos % of Long'], ascending=[True, False])
        result = df_long_avg.groupby('entry_date').head(count).groupby('entry_date')['pos % of Long'].mean().reset_index()
        avg_long_position = result['pos % of Long'].mean()

        df_nav_avg = df_position_long.sort_values(by=['entry_date', 'pos % of NAV'], ascending=[True, False])
        result = df_nav_avg.groupby('entry_date').head(count).groupby('entry_date')['pos % of NAV'].mean().reset_index()
        avg_long_position_nav = result['pos % of NAV'].mean()

        df_short_avg = df_position_short.sort_values(by=['entry_date', 'pos % of Long'], ascending=[True, True])
        result = df_short_avg.groupby('entry_date').head(count).groupby('entry_date')['pos % of Long'].mean().reset_index()
        avg_short_position = result['pos % of Long'].mean()

        df_nav_avg = df_position_short.sort_values(by=['entry_date', 'pos % of NAV'], ascending=[True, True])
        result = df_nav_avg.groupby('entry_date').head(count).groupby('entry_date')['pos % of NAV'].mean().reset_index()
        avg_short_position_nav = result['pos % of NAV'].mean()

        df_avg_size = df_avg_size._append(pd.DataFrame({'count': count, 'Avg Long Pos vs Long': avg_long_position,
                                                        'Avg Short Pos vs Long': -avg_short_position,
                                                        'Avg Long Pos vs NAV - Low Vol': avg_long_position_nav,
                                                        'Avg Short Pos vs NAV - Low Vol': -avg_short_position_nav,
                                                        'Avg Long Pos vs NAV - Alto': avg_long_position_nav*2,
                                                        'Avg Short Pos vs NAV - Alto': -avg_short_position_nav*2
                                                        }, index=[0]), ignore_index=True)

    # in df_position_long, rename 'pos % of NAV' to 'pos % of NAV - Low vol'
    df_position_long = df_position_long.rename(columns={'pos % of NAV': 'pos % of NAV - Low vol'})
    df_position_long['pos % of NAV - Alto'] = df_position_long['pos % of NAV - Low vol']*2
    df_position_short = df_position_short.rename(columns={'pos % of NAV': 'pos % of NAV - Low vol'})
    df_position_short['pos % of NAV - Alto'] = df_position_short['pos % of NAV - Low vol']*2

    # convert entry_date to date
    df_position_long['entry_date'] = df_position_long['entry_date'].dt.date
    df_position_short['entry_date'] = df_position_short['entry_date'].dt.date

    excel_writer = pd.ExcelWriter('Excel/Single Position Size.xlsx', engine='openpyxl')
    df_position_long.to_excel(excel_writer, sheet_name='Position Long - All', startrow=0, index=False, header=True)
    df_position_short.to_excel(excel_writer, sheet_name='Position Short - All', startrow=0, index=False, header=True)
    df_avg_size.to_excel(excel_writer, sheet_name='Avg Position Size', startrow=0, index=False, header=True)
    excel_writer._save()

    workbook = load_workbook('Excel/Single Position Size.xlsx')
    sheet_all_names = ['Position Long - All', 'Position Short - All']

    for sheet in sheet_all_names:
        sheet = workbook[sheet]
        for col in range(ord('F'), ord('H') + 1):  # Loop through columns B to G
            column_letter = chr(col)
            for cell in sheet[column_letter]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

    sheet = workbook['Avg Position Size']
    sheet.column_dimensions['A'].number_format = 'General'

    # for A to G column, format %

    for col in range(ord('B'), ord('G') + 1):  # Loop through columns B to G
        column_letter = chr(col)
        for cell in sheet[column_letter]:
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    fill = PatternFill(start_color='66b8cc', end_color='66b8cc', fill_type='solid')

    for col in range(1, 8):
        cell = sheet.cell(row=1, column=col)
        cell.fill = fill

    sheet_list = workbook.sheetnames
    for sheet_name in sheet_list:
        sheet = workbook[sheet_name]
        # column autofit
        for col in sheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1  # Adding some buffer and a conversion factor
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    workbook.move_sheet('Avg Position Size', offset=-2)

    workbook.save('Excel/Single Position Size.xlsx')


if __name__ == '__main__':
    get_position_size()
