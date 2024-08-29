import pandas as pd
from models import engine, config_class
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date
import numpy as np

"""
Get the top X contributors/detractors in cash equity in absolute pnl (% of AUM) or by Alpha per month and per year
"""

def get_top_contributors(start_date, end_date, my_type, my_num=5):
    my_sql = f"""SELECT T2.ticker,T1.entry_date,T1.{my_type} FROM position T1
JOIN Product T2 ON T1.product_id = T2.id WHERE prod_type = 'Cash' AND entry_date>='{start_date}'
 AND entry_date<='{end_date}' AND parent_fund_id=1
order by entry_date"""

    df_position = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    my_sql = """SELECT entry_date,amount*1000000 as amount FROM aum WHERE type='leveraged' and fund_id=4;"""
    df_aum = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    my_sql ="""SELECT entry_date,long_usd FROM alpha_summary WHERE parent_fund_id=1 order by entry_date;"""
    df_long = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    # merge
    df_temp = pd.merge(df_position, df_aum, on='entry_date', how='left')
    df_temp = pd.merge(df_temp, df_long, on='entry_date', how='left')
    # fill with previous value
    df_temp['amount'] = df_temp['amount'].fillna(method='ffill')

    if my_type == 'pnl_usd':
        df_temp['perf'] = 2 * df_temp[my_type] / df_temp['amount']
    else:
        df_temp['perf'] = 2 * df_temp[my_type] / df_temp['long_usd']

    df_temp['month_year'] = df_temp['entry_date'].dt.strftime('%Y-%m')
    df = df_temp.groupby(['month_year', 'ticker'])['perf'].sum().reset_index()

    # sort by month_year and perf
    df = df.sort_values(by=['month_year', 'perf'], ascending=[True, False])

    # get all unique month_year
    month_year = df['month_year'].unique()
    # get top X contributors
    df_top_month = pd.DataFrame()
    df_bottom_month = pd.DataFrame()
    for my_month_year in month_year:
        df_current_month = df[df['month_year'] == my_month_year]

        df_top_X = df_current_month.head(my_num)
        df_top_month[my_month_year] = df_top_X['ticker'].values

        df_bottom_X = df_current_month.tail(my_num)
        df_bottom_X = df_bottom_X.sort_values(by='perf', ascending=True)
        df_bottom_month[my_month_year] = df_bottom_X['ticker'].values

    df_temp['year'] = df_temp['entry_date'].dt.strftime('%Y')
    df = df_temp.groupby(['year', 'ticker'])['perf'].sum().reset_index()
    df = df.sort_values(by=['year', 'perf'], ascending=[True, False])
    df_top_year_ticker = pd.DataFrame()
    df_bottom_year_ticker = pd.DataFrame()

    df_top_year_bp = pd.DataFrame()
    df_bottom_year_bp = pd.DataFrame()

    year = df['year'].unique()
    for my_year in year:
        df_top_X = df[df['year'] == my_year].head(my_num)
        df_top_year_ticker[my_year] = df_top_X['ticker'].values
        df_top_year_bp[my_year] = np.round(df_top_X['perf'].values*10000, 1)

        df_bottom_X = df[df['year'] == my_year].tail(my_num)
        df_bottom_X = df_bottom_X.sort_values(by='perf', ascending=True)
        df_bottom_year_ticker[my_year] = df_bottom_X['ticker'].values
        df_bottom_year_bp[my_year] = np.round(df_bottom_X['perf'].values*10000, 1)

    if my_type == 'pnl_usd':
        file_name = f'Excel/Return Based Contributors-Detractors - Top {my_num}.xlsx'
    else:
        file_name = f'Excel/Alpha Based Contributors-Detractors - Top {my_num}.xlsx'
    # add incremental index starting at 1
    df_top_month.index = df_top_month.index + 1
    df_bottom_month.index = df_bottom_month.index + 1
    df_top_year_ticker.index = df_top_year_ticker.index + 1
    df_bottom_year_ticker.index = df_bottom_year_ticker.index + 1
    df_top_year_bp.index = df_top_year_bp.index + 1
    df_bottom_year_bp.index = df_bottom_year_bp.index + 1

    excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')
    df_top_month.to_excel(excel_writer, sheet_name='Sheet1', startrow=1, index=True, header=True)
    df_bottom_month.to_excel(excel_writer, sheet_name='Sheet1', startrow=my_num+5, index=True, header=True)
    df_top_year_ticker.to_excel(excel_writer, sheet_name='Sheet1', startrow=my_num*2+9, index=True, header=True)
    df_bottom_year_ticker.to_excel(excel_writer, sheet_name='Sheet1', startrow=my_num*3+13, index=True, header=True)
    df_top_year_bp.to_excel(excel_writer, sheet_name='Sheet1', startrow=my_num*4+17, index=True, header=True)
    df_bottom_year_bp.to_excel(excel_writer, sheet_name='Sheet1', startrow=my_num*5+21, index=True, header=True)

    excel_writer._save()

    workbook = load_workbook(file_name)
    sheet = workbook.active

    # Merge cells and apply styles

    format_table = [
        ['A1:D1', 'A1', f'Top 5 Contributors by month', "66b8cc"],
        [f'A{my_num + 5}:D{my_num + 5}', f'A{my_num + 5}', f'Top {my_num} Detractors by month', "eb7373"],
        [f'A{my_num * 2 + 9}:D{my_num * 2 + 9}', f'A{my_num * 2 + 9}', f'Top {my_num} Contributors by year', "66b8cc"],
        [f'A{my_num * 3 + 13}:D{my_num * 3 + 13}', f'A{my_num * 3 + 13}', f'Top {my_num} Detractors by year', "eb7373"],
        [f'A{my_num * 4 + 17}:D{my_num * 4 + 17}', f'A{my_num * 4 + 17}', f'Top {my_num} Contributors BP by year', "66b8cc"],
        [f'A{my_num * 5 + 21}:D{my_num * 5 + 21}', f'A{my_num * 5 + 21}', f'Top {my_num} Detractors BP by year', "eb7373"]
        ]

    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1  # Adding some buffer and a conversion factor
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    for format_list in format_table:
        sheet.merge_cells(format_list[0])
        merged_cell = sheet[format_list[1]]
        merged_cell.value = format_list[2]
        merged_cell.font = Font(bold=True)
        merged_cell.fill = PatternFill(start_color=format_list[3], end_color=format_list[3], fill_type="solid")

    workbook.save(file_name)


if __name__ == '__main__':

    start_date = date(2019, 4, 1)
    end_date = date(2024, 5, 31)

    get_top_contributors(start_date, end_date, 'alpha_usd', 5)
    get_top_contributors(start_date, end_date, 'pnl_usd', 5)
