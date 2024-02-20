import pandas as pd
from models import engine, config_class
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

"""
Asset_class: Get the % of asset class traded per year. 
We take the net volume traded by day by product (long+short in absolute value). 
We divide it by Cash Equity, CFD, Equity index and the rest
"""

def get_asset_class(start_date, end_date):
    my_sql = f"""SELECT T1.trade_date,T2.ticker,T2.prod_type,T4.name as country,T4.continent,
    T5.name as sector,T6.generic_future as security,abs(sum(T1.notional_usd)) as notional_usd 
    FROM trade T1 JOIN product T2 on T1.product_id=T2.id
    LEFT JOIN exchange T3 on T2.exchange_id=T3.id JOIN country T4 on T3.country_id=T4.id
    LEFT JOIN industry_sector T5 on T2.industry_sector_id=T5.id 
    LEFT JOIN security T6 on T2.security_id=T6.id WHERE parent_fund_id=1 and trade_date>='{start_date}'
    and trade_date<='{end_date}' and T2.ticker <>'EDZ2 CME'
    GROUP BY trade_date,T2.ticker,T2.prod_type,country,continent,sector order by T2.ticker,trade_date;"""
    df_trade = pd.read_sql(my_sql, con=engine, parse_dates=['trade_date'])

    # add column asset_class
    df_trade['asset_class'] = 'Other'

    # if prod_type is Cash, and country in ('united States, Canada, Switzerland'), then asset_class is 'Cash'
    df_trade.loc[(df_trade['prod_type'] == 'Cash') &
                 (df_trade['country'].isin(['United States', 'Canada', 'Switzerland'])), 'asset_class'] = 'Cash Equity'
    # if prod_type is cash and country not in ('united States, Canada, Switzerland'), then asset_class is 'CFD'
    df_trade.loc[(df_trade['prod_type'] == 'Cash') &
                 (~df_trade['country'].isin(['United States', 'Canada', 'Switzerland'])), 'asset_class'] = 'CFD'
    # if ticker in ('SXO1 EUX', 'ES1 CME'), then asset_class is 'Equity Index'
    df_trade.loc[df_trade['security'].isin(['SXO1 EUX', 'ES1 CME']), 'asset_class'] = 'Equity Index'
    df_trade['notional_usd'] = df_trade['notional_usd'].fillna(0)
    # pivot table with asset_class as columns, year as index, and sum of notional_usd as values
    df_asset_class = pd.pivot_table(df_trade, values='notional_usd', index=df_trade['trade_date'].dt.year,
                                    columns='asset_class', aggfunc='sum', fill_value=0)

    df_asset_class = df_asset_class.div(df_asset_class.sum(axis=1), axis=0).round(2)
    # into excel
    filename = 'Excel\Asset Class per year.xlsx'
    df_asset_class.to_excel(filename, index=True)
    workbook = load_workbook(filename)
    worksheet = workbook.active
    for col_num in range(2, 6):
        column_letter = worksheet.cell(row=1, column=col_num).column_letter
        for cell in worksheet[column_letter]:
            cell.number_format = numbers.FORMAT_PERCENTAGE

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1) * 1.1  # Adding some buffer and a conversion factor
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    cell = worksheet.cell(row=1, column=1)
    cell.value = ""

    workbook.save(filename)


if __name__ == '__main__':
    start_date = '2019-04-01'
    end_date = '2023-12-31'
    get_asset_class(start_date, end_date)



