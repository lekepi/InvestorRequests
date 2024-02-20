import pandas as pd
from models import engine
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers

'''
Calculate in excel the exposure or attribution.
You can decide the list of classification (Sector, Country, Continent, Market cap...) that will be one tab in the excel file.
You can decide the list of numerator (Long, Short, Net) that will be extra columns in the excel file.
You can the base / denominator (NAV, long, Short, gross, net).
You can decide the start and end date.
You can decide if the exposure is leveraged or not.
You can decide if the pnl intraday from trade is included or not.
'''


def get_monthly_data(calcul_type, classification_list, numerator_list,
                     denominator, start_date, end_date, is_leveraged, is_close_pnl):

    # get list of start date of the month
    my_sql = f"""SELECT MIN(entry_date) AS first_date FROM position WHERE entry_date>='{start_date}' 
        and entry_date<='{end_date}' GROUP BY YEAR(entry_date), MONTH(entry_date) order by entry_date"""
    df_date = pd.read_sql(my_sql, con=engine, parse_dates=['first_date'])
    first_date_list = df_date['first_date'].tolist()
    first_date_str = ','.join([f"'{first_date}'" for first_date in first_date_list])

    # df exposure
    my_sql = f"""SELECT T1.entry_date,T2.ticker,T2.prod_type,T4.name as Country,T4.continent as Continent,
    T5.name as Sector,sum(T1.mkt_value_usd) as notional_usd FROM position T1 JOIN product T2 on T1.product_id=T2.id
    LEFT JOIN exchange T3 on T2.exchange_id=T3.id LEFT JOIN country T4 on T3.country_id=T4.id
    LEFT JOIN industry_group_gics T5 on T2.industry_group_gics_id=T5.id WHERE parent_fund_id=1 and entry_date in ({first_date_str}) and (T2.prod_type='Cash' or T2.ticker in ('ES1 CME', 'SXO1 EUX'))
    GROUP BY entry_date,T2.ticker,T2.prod_type,Country,continent,sector order by T2.ticker,entry_date;"""
    df_exposure = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
    df_exposure['month_year'] = df_exposure['entry_date'].dt.strftime('%Y-%m')
    df_exposure = df_exposure.drop(columns=['entry_date'])

    # get the denominator: long, short, net, gross, nav
    if denominator == 'Long':
        df_denominator = df_exposure[df_exposure['notional_usd'] >= 0]
        df_denominator = df_denominator.groupby('month_year').agg({'notional_usd': 'sum'}).reset_index()
        df_denominator.rename(columns={'notional_usd': 'denominator'}, inplace=True)
    elif denominator == 'Short':
        df_denominator = df_exposure[df_exposure['notional_usd'] < 0]
        df_denominator = df_denominator.groupby('month_year').agg({'notional_usd': 'sum'}).reset_index()
        df_denominator.rename(columns={'notional_usd': 'denominator'}, inplace=True)
    elif denominator == 'Gross':
        # absolute value of notional_usd
        df_denominator = df_exposure.copy()
        df_denominator['notional_usd'] = df_denominator['notional_usd'].abs()
        df_denominator = df_denominator.groupby('month_year').agg({'notional_usd': 'sum'}).reset_index()
        df_denominator.rename(columns={'notional_usd': 'denominator'}, inplace=True)
    elif denominator == 'Net':
        df_denominator = df_exposure.groupby('month_year').agg({'notional_usd': 'sum'}).reset_index()
        df_denominator.rename(columns={'notional_usd': 'denominator'}, inplace=True)
    elif denominator == 'Nav':
        my_sql = f"""Select entry_date,amount*1000000 as denominator from aum where entry_date>='2019-04-01' 
                and entry_date>='{start_date}' and entry_date<='{end_date}' and type='leveraged' order by entry_date"""
        df_denominator = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        df_denominator['month_year'] = df_denominator['entry_date'].dt.strftime('%Y-%m')
        df_denominator.drop(columns=['entry_date'], inplace=True)

    df_exposure.sort_values(by='month_year', ascending=False, inplace=True)
    df_exposure['notional_usd'] = df_exposure['notional_usd'].fillna(0)
    df_exposure = df_exposure.merge(df_denominator, on='month_year', how='left')
    df_exposure['Exposure'] = df_exposure['notional_usd'] / df_exposure['denominator']
    if calcul_type == 'Exposure':
        df = df_exposure.copy()
    else:  # calcul_type != 'Exposure'
        # df_pnl
        my_sql = f"""SELECT min(T1.entry_date) as entry_date,T2.ticker,T2.prod_type,T4.name as Country,T4.continent as Continent,
        T5.name as Sector,sum(T1.pnl_usd) as pnl_usd, sum(T1.alpha_usd) as alpha_usd FROM position T1 JOIN product T2 on T1.product_id=T2.id
        LEFT JOIN exchange T3 on T2.exchange_id=T3.id LEFT JOIN country T4 on T3.country_id=T4.id
        LEFT JOIN industry_group_gics T5 on T2.industry_group_gics_id=T5.id WHERE parent_fund_id=1
        and entry_date>='{start_date}' and entry_date<='{end_date}' and (T2.prod_type='Cash' or T2.ticker in ('ES1 CME', 'SXO1 EUX'))
        GROUP BY YEAR(entry_date),MONTH(entry_date),T2.ticker,T2.prod_type,Country,continent,sector order by T2.ticker,entry_date;"""
        df_pnl = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        df_pnl['month_year'] = df_pnl['entry_date'].dt.strftime('%Y-%m')
        # remove entry_date
        df_pnl = df_pnl.drop(columns=['entry_date'])

        if is_close_pnl:
            my_sql = f"""SELECT min(T1.trade_date) as entry_date,
                CASE WHEN T3.generic_future IS NOT NULL THEN T3.generic_future ELSE T2.ticker END as ticker,sum(T1.pnl_close) as pnl_close FROM trade T1 
                JOIN product T2 on T1.product_id=T2.id LEFT JOIN security T3 on T2.security_id=T3.id WHERE T1.parent_fund_id=1 and trade_date>='2019-04-01'
                and trade_date>='{start_date}' and trade_date<='{end_date}' and (T2.prod_type in ('Cash', 'Future') or T3.generic_future in 
                ('ES1 CME', 'SXO1 EUX')) GROUP BY YEAR(trade_date),MONTH(trade_date),ticker order by ticker,T1.trade_date;"""
            df_pnl_close = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
            df_pnl_close['month_year'] = df_pnl_close['entry_date'].dt.strftime('%Y-%m')
            # remove entry_date
            df_pnl_close = df_pnl_close.drop(columns=['entry_date'])
            df_pnl = df_pnl.merge(df_pnl_close, on=['month_year', 'ticker'], how='left')
            df_pnl['pnl_close'] = df_pnl['pnl_close'].fillna(0)
            df_pnl['pnl_usd'] = df_pnl['pnl_usd'] + df_pnl['pnl_close']
            df_pnl['alpha_usd'] = df_pnl['alpha_usd'] + df_pnl['pnl_close']

        df = df_exposure.merge(df_pnl, on=['month_year', 'ticker', 'prod_type', 'Country', 'Continent', 'Sector'], how='left')

        df['pnl_usd'] = df['pnl_usd'].fillna(0)
        df['PnL'] = df['pnl_usd'] / df['denominator']
        df['Alpha'] = df['alpha_usd'] / df['denominator']

    if 'MarketCap' in classification_list:
        my_sql = f"""SELECT T1.entry_date,T2.ticker,CASE WHEN T3.market_cap<3000 then '0-3Bn' WHEN T3.market_cap<10000 THEN '3-10Bn' else '>10Bn' END as marketCap
        FROM position T1 JOIN product T2 on T1.product_id=T2.id JOIN product_market_cap T3 on T1.product_id=T3.product_id and T1.entry_date=T3.entry_date and T3.type='Monthly'
        WHERE T1.entry_date in ({first_date_str})  and T2.prod_type='Cash' and T1.parent_fund_id=1
        GROUP by T1.entry_date,T2.ticker;"""
        df_market_cap = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])
        df_market_cap['month_year'] = df_market_cap['entry_date'].dt.strftime('%Y-%m')
        df_market_cap = df_market_cap.drop(columns=['entry_date'])
        df = df.merge(df_market_cap, on=['month_year', 'ticker'], how='left')
        df['MarketCap'] = df.groupby('ticker')['marketCap'].fillna(method='ffill')
        df['MarketCap'] = df.apply(lambda x: 'Index' if x['ticker'] in ('ES1 CME', 'SXO1 EUX') else x['MarketCap'], axis=1)

    df['notional_usd'] = df['notional_usd'].fillna(0)
    df['Sector'] = df.apply(lambda x: 'Index' if x['ticker'] in ('ES1 CME', 'SXO1 EUX') else x['Sector'], axis=1)
    df['Country'] = df.apply(lambda x: 'S&P500' if x['ticker'] in ('ES1 CME') else x['Country'], axis=1)
    df['Country'] = df.apply(lambda x: 'Stoxx 600' if x['ticker'] in ('SXO1 EUX') else x['Country'], axis=1)

    df.sort_values(by='month_year', ascending=False, inplace=True)

    if is_leveraged and denominator == 'Nav':
        df['Exposure'] = df['Exposure'] * 2
        if 'PnL' in df.columns:
            df['PnL'] = df['PnL'] * 2
        if 'Alpha' in df.columns:
            df['Alpha'] = df['Alpha'] * 2

    result_list = []
    for classification in classification_list:
        df_result = pd.DataFrame()
        numerator_len_list = []
        for numerator in numerator_list:
            df_temp = df.copy()
            if numerator == 'Long':
                df_temp = df_temp[df_temp['notional_usd'] >= 0]
            elif numerator == 'Short':
                df_temp = df_temp[df_temp['notional_usd'] < 0]

            if (numerator == 'Long' and denominator == 'Short') or (numerator == 'Short' and denominator == 'Long'):
                df_temp['Exposure'] = df_temp['Exposure'] * -1

            df_temp = df_temp.groupby(['month_year', classification]).agg({f'{calcul_type}': 'sum'}).reset_index()
            df_pivot = df_temp.pivot(index='month_year', columns=classification, values=calcul_type).fillna(0)
            if 'Index' in df_pivot.columns:
                df_pivot = df_pivot[[col for col in df_pivot.columns if col != 'Index'] + ['Index']]
            if 'S&P500' in df_pivot.columns:
                df_pivot = df_pivot[[col for col in df_pivot.columns if col != 'S&P500'] + ['S&P500']]
            if 'Stoxx 600' in df_pivot.columns:
                df_pivot = df_pivot[[col for col in df_pivot.columns if col != 'Stoxx 600'] + ['Stoxx 600']]
            if len(numerator_list) > 1:
                # add numerator to the column name
                df_pivot.columns = [f'{col} {numerator}' for col in df_pivot.columns]
            numerator_len_list.append(len(df_pivot.columns))
            if df_result.empty:
                df_result = df_pivot
            else:
                df_result = df_result.merge(df_pivot, on='month_year', how='outer')
        df_result = df_result.sort_values(by='month_year', ascending=False)
        result_list.append([classification, df_result, numerator_len_list])

    # create the excel file

    if calcul_type == 'PnL':
        calcul_type_string = 'Attribution'
    else:
        calcul_type_string = calcul_type

    if len(numerator_list) == 1:
        numerator_str = numerator_list[0] + ' '
    else:
        numerator_str = ' '

    file_name = f'Excel/Monthly {numerator_str}{calcul_type_string} Over {denominator}.xlsx'
    with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
        for result in result_list:
            result[1].to_excel(writer, sheet_name=result[0], header=True, index=True)

    # complete formatting

    fill1 = PatternFill(start_color='66b8cc', end_color='66b8cc', fill_type='solid')
    fill2 = PatternFill(start_color='f08c30', end_color='f08c30', fill_type='solid')

    workbook = load_workbook(file_name)
    for result in result_list:
        worksheet = workbook[result[0]]
        numerator_len_list = result[2]

        df_temp = result[1]  # get the df
        num_cols = len(df_temp.columns)
        # header color

        # for col_num in range(1, num_cols + 2):
        #    worksheet.cell(row=1, column=col_num).fill = fill1
        current_fill = fill1
        col_index = 1
        for num_cols_same_color in numerator_len_list:
            for _ in range(num_cols_same_color):
                worksheet.cell(row=1, column=col_index+1).fill = current_fill
                col_index += 1
            # Switch fill
            current_fill = fill2 if current_fill == fill1 else fill1

            # Format columns 2 to the last as percentages
        # % format
        for col_num in range(2, num_cols + 2):
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            for cell in worksheet[column_letter]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

        for i, col in enumerate(worksheet.columns):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = ((max_length+2) * 1.2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Add freeze pane to the second row
        worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    workbook.save(file_name)


if __name__ == '__main__':

    # calcul_type = 'Exposure', 'PnL', 'Alpha'
    # classification_list = ['Continent', 'Sector', 'MarketCap']
    # numerator_list = ['Long'], ['Short'], ['Net'], ['Gross']
    # denominator = 'Long', 'Nav', 'Gross', 'Net'

    start_date = date(2019, 4, 1)
    end_date = date(2023, 12, 31)

    # exposure vs Short only
    get_monthly_data(calcul_type='Exposure',
                     classification_list=['Continent', 'Sector', 'MarketCap'],
                     numerator_list=['Short'],
                     denominator='Short',
                     start_date=start_date,
                     end_date=end_date,
                     is_leveraged=False,
                     is_close_pnl=True)

    # exposure vs Short / Long
    get_monthly_data(calcul_type='Exposure',
                     classification_list=['Continent', 'Sector', 'MarketCap'],
                     numerator_list=['Short'],
                     denominator='Long',
                     start_date=start_date,
                     end_date=end_date,
                     is_leveraged=False,
                     is_close_pnl=True)


    # exposure vs Nav - Leveraged
    get_monthly_data(calcul_type='Exposure',
                     classification_list=['Continent', 'Sector', 'MarketCap'],
                     numerator_list=['Net', 'Long', 'Short'],
                     denominator='Nav',
                     start_date=start_date,
                     end_date=end_date,
                     is_leveraged=True,
                     is_close_pnl=True)

    # exposure vs Long only
    get_monthly_data(calcul_type='Exposure',
                     classification_list=['Continent', 'Sector', 'MarketCap'],
                     numerator_list=['Long'],
                     denominator='Long',
                     start_date=start_date,
                     end_date=end_date,
                     is_leveraged=False,
                     is_close_pnl=True)

    # pnl vs Nav
    get_monthly_data(calcul_type='PnL',
                     classification_list=['Continent', 'Sector', 'MarketCap'],
                     numerator_list=['Net', 'Long', 'Short'],
                     denominator='Nav',
                     start_date=start_date,
                     end_date=end_date,
                     is_leveraged=True,
                     is_close_pnl=True)
