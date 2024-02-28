import pandas as pd
from models import engine, Aum, session
from datetime import timedelta, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers


def get_exposure_last(classification_list, numerator_list, denominator, entry_date, is_leveraged):
    # df exposure
    my_sql = f"""SELECT T2.ticker,T2.prod_type,T4.name as Country,T4.continent as Continent,
       T5.name as Sector,sum(T1.mkt_value_usd) as notional_usd FROM position T1 JOIN product T2 on T1.product_id=T2.id
       LEFT JOIN exchange T3 on T2.exchange_id=T3.id LEFT JOIN country T4 on T3.country_id=T4.id
       LEFT JOIN industry_group_gics T5 on T2.industry_group_gics_id=T5.id WHERE parent_fund_id=1 
       and entry_date='{entry_date}' and (T2.prod_type='Cash' or T2.ticker in ('ES1 CME', 'SXO1 EUX'))
       GROUP BY entry_date,T2.ticker,T2.prod_type,Country,continent,sector order by T2.ticker,entry_date;"""
    df_exposure = pd.read_sql(my_sql, con=engine, parse_dates=['entry_date'])

    # get the denominator: long, short, net, gross, nav

    df_temp = df_exposure[df_exposure['notional_usd'] > 0]
    long_value = df_temp['notional_usd'].sum()
    df_temp = df_exposure[df_exposure['notional_usd'] < 0]
    short_value = -df_temp['notional_usd'].sum()
    gross_value = long_value + short_value
    net_value = df_exposure['notional_usd'].sum()

    if denominator == 'Long':
        denominator_value = long_value
    elif denominator == 'Short':
        denominator_value = short_value
    elif denominator == 'Gross':
        denominator_value = gross_value
    elif denominator == 'Net':
        denominator_value = net_value
    elif denominator == 'Nav':
        # first day of the month from entry_date
        start_date = date(entry_date.year, entry_date.month, 1)
        #if WE go to next monday
        if start_date.weekday() == 5:
            start_date = start_date + timedelta(days=2)
        elif start_date.weekday() == 6:
            start_date = start_date + timedelta(days=1)

        aum = session.query(Aum).filter(Aum.entry_date == start_date, Aum.type == 'leveraged').all()
        if aum:
            denominator_value = aum[0].amount * 1000000

    df_exposure['Exposure'] = df_exposure['notional_usd'] / denominator_value

    df = df_exposure.copy()

    if is_leveraged and denominator == 'Nav':
        df['Exposure'] = df['Exposure'] * 2

    df['Sector'] = df.apply(lambda x: 'Index' if x['ticker'] in ('ES1 CME', 'SXO1 EUX') else x['Sector'], axis=1)
    df['Country'] = df.apply(lambda x: 'S&P500' if x['ticker'] in ('ES1 CME') else x['Country'], axis=1)
    df['Country'] = df.apply(lambda x: 'Stoxx 600' if x['ticker'] in ('SXO1 EUX') else x['Country'], axis=1)

    result_list = []
    for classification in classification_list:
        df_result = pd.DataFrame()
        for numerator in numerator_list:
            df_temp = df.copy()
            if (numerator == 'Long' and denominator == 'Short') or (numerator == 'Short' and denominator == 'Long'):
                df_temp['Exposure'] = df_temp['Exposure'] * -1
            if numerator == 'Long':
                df_temp = df_temp[df_temp['notional_usd'] >= 0]
            elif numerator == 'Short':
                df_temp = df_temp[df_temp['notional_usd'] < 0]
            elif numerator == 'Gross':
                df_temp['Exposure'] *= df_temp['Exposure'].apply(lambda x: 1 if x >= 0 else -1)

            # group by classification, no Pivot !
            df_group = df_temp.groupby([classification])['Exposure'].sum().reset_index()
            # rename column exposure to classification
            df_group = df_group.rename(columns={'Exposure': numerator})

            if df_result.empty:
                df_result = df_group
            else:
                df_result = pd.merge(df_result, df_group, on=classification, how='outer')
        # fillna in 0
        df_result = df_result.fillna(0)
        # sort by long desc
        df_result = df_result.sort_values(by=numerator_list[0], ascending=False)
        result_list.append((classification, df_result))

    # create the excel file

    if len(numerator_list) == 1:
        numerator_str = numerator_list[0] + ' '
    else:
        numerator_str = ' '

    file_name = f'Excel/Last {numerator_str}Exposure Over {denominator}.xlsx'
    with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
        for result in result_list:
            result[1].to_excel(writer, sheet_name=result[0], header=True, index=True)

    workbook = load_workbook(file_name)
    for result in result_list:
        worksheet = workbook[result[0]]

        df_temp = result[1]  # get the df
        num_cols = len(df_temp.columns)

        for col_num in range(2, num_cols + 2):
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            for cell in worksheet[column_letter]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

        for i, col in enumerate(worksheet.columns):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = ((max_length + 2) * 1.2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Add freeze pane to the second row
        worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    workbook.save(file_name)


if __name__ == '__main__':

    entry_date = date(2024, 1, 1)
    get_exposure_last(classification_list=['Country', 'Sector'],
                      numerator_list=['Long', 'Short', 'Gross', 'Net'],
                     denominator='Nav', entry_date=entry_date, is_leveraged=True)
